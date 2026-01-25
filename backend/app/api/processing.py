import os
import json
import sys
from fastapi import APIRouter, HTTPException, BackgroundTasks, File, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
import asyncio
from typing import List, Dict, Optional
import uuid
from app.services.document_processor import DocumentProcessor
from app.services.mismatch_detector import MismatchDetector
from app.services.excel_generator import ExcelGenerator
from app.config import UPLOAD_DIR
from openpyxl import load_workbook
import tempfile

router = APIRouter()

# In-memory storage for processing jobs (in production, use a database)
processing_jobs: Dict[str, Dict] = {}

class ProcessingSession:
    """Manages a processing session for documents"""
    
    def __init__(self, session_id: str, client_name: str, month: str):
        self.session_id = session_id
        self.client_name = client_name
        self.month = month
        self.status = "initialized"
        self.progress = 0
        self.extracted_invoices = []
        self.gstr2b_data = None
        self.mismatch_results = None
        self.excel_data = None
        self.error = None
    
    def to_dict(self):
        return {
            "session_id": self.session_id,
            "client_name": self.client_name,
            "month": self.month,
            "status": self.status,
            "progress": self.progress,
            "extracted_invoices": self.extracted_invoices,
            "gstr2b_data": self.gstr2b_data,
            "mismatch_results": self.mismatch_results,
            "excel_data": self.excel_data,
            "error": self.error
        }


def _parse_gstr2b_excel(file_path: str) -> Dict:
    """
    Parse GSTR2B Excel file and extract invoice data
    """
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        invoices = []
        headers = []
        
        # Get headers from first row
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip())
            else:
                headers.append("")
        
        # Map common GSTR2B column names to our expected fields
        column_mapping = {
            'invoice_no': ['invoice_no', 'invoice number', 'invoiceno', 'inv_no'],
            'invoice_date': ['invoice_date', 'invoice date', 'invoicedate', 'inv_date'],
            'supplier_gstin': ['supplier_gstin', 'gstin', 'vendor_gstin', 'supplier gstin'],
            'taxable_value': ['taxable_value', 'taxable value', 'amount', 'invoice_amount'],
            'cgst': ['cgst', 'cgst amount', 'central gst'],
            'sgst': ['sgst', 'sgst amount', 'state gst'],
            'igst': ['igst', 'igst amount', 'integrated gst'],
            'total_amount': ['total_amount', 'total amount', 'total', 'grand total'],
            'gst_rate': ['gst_rate', 'gst rate', 'rate']
        }
        
        # Find column indices
        col_indices = {}
        for field, possible_names in column_mapping.items():
            for idx, header in enumerate(headers):
                if header in possible_names:
                    col_indices[field] = idx
                    break
        
        # Extract invoice data
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            if all(cell.value is None for cell in row):
                continue
            
            invoice = {}
            for field, col_idx in col_indices.items():
                if col_idx < len(row):
                    value = row[col_idx].value
                    if value is not None:
                        if field in ['taxable_value', 'cgst', 'sgst', 'igst', 'total_amount', 'gst_rate']:
                            try:
                                invoice[field] = float(value)
                            except (ValueError, TypeError):
                                invoice[field] = 0
                        else:
                            invoice[field] = str(value).strip()
            
            if invoice:
                invoices.append(invoice)
        
        if not invoices:
            return {"invoices": []}
        
        # Extract GSTIN from supplier_gstin if available
        gstin = None
        if invoices and 'supplier_gstin' in invoices[0]:
            gstin = invoices[0]['supplier_gstin']
        
        return {
            "gstin": gstin,
            "period": None,  # Period should be provided by the user or extracted from file metadata
            "invoices": invoices
        }
    
    except Exception as e:
        print(f"Error parsing GSTR2B Excel: {str(e)}", file=sys.stderr)
        return {"invoices": []}


@router.post("/process")
async def process_documents(client_name: str, month: str, background_tasks: BackgroundTasks):
    """
    Initiate document processing for uploaded files
    Returns a session ID for tracking progress
    """
    print(f"\n[PROCESS] Starting process endpoint: client={client_name}, month={month}", file=sys.stderr)
    
    try:
        # Create processing session
        session_id = str(uuid.uuid4())
        session = ProcessingSession(session_id, client_name, month)
        processing_jobs[session_id] = session
        print(f"[PROCESS] Created session: {session_id}", file=sys.stderr)
        
        # Get file paths
        client_path = os.path.join(UPLOAD_DIR, client_name, month)
        print(f"[PROCESS] Looking for files in: {client_path}", file=sys.stderr)
        
        if not os.path.exists(client_path):
            print(f"[PROCESS] ERROR: Directory not found: {client_path}", file=sys.stderr)
            raise HTTPException(status_code=400, detail="Upload directory not found")
        
        # Get all files in the directory (including nested subdirectories)
        file_paths = []
        for root, dirs, files in os.walk(client_path):
            for f in files:
                # Only include document files (PDF, images, etc.)
                if f.lower().endswith(('.pdf', '.png', '.jpg', '.jpeg', '.jpg', '.tiff', '.tif')):
                    file_path = os.path.join(root, f)
                    file_paths.append(file_path)
                    print(f"[PROCESS] Found file: {file_path}", file=sys.stderr)
        
        all_items = os.listdir(client_path)
        print(f"[PROCESS] Items in directory: {all_items}", file=sys.stderr)
        print(f"[PROCESS] Found {len(file_paths)} document files: {[os.path.basename(f) for f in file_paths]}", file=sys.stderr)
        
        if not file_paths:
            print(f"[PROCESS] ERROR: No document files found in {client_path}", file=sys.stderr)
            raise HTTPException(status_code=400, detail="No document files found in upload directory. Supported formats: PDF, PNG, JPG, JPEG, TIFF")
        
        # Start background processing
        print(f"[PROCESS] Starting background task for {len(file_paths)} files", file=sys.stderr)
        background_tasks.add_task(
            process_documents_background,
            session_id,
            file_paths
        )
        print(f"[PROCESS] ✓ Background task queued for session {session_id}", file=sys.stderr)
        
        return {
            "status": "processing_started",
            "session_id": session_id,
            "client_name": client_name,
            "month": month,
            "file_count": len(file_paths)
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


async def process_documents_background(session_id: str, file_paths: List[str]):
    """Background task for processing documents"""
    session = processing_jobs[session_id]
    print(f"\n[BACKGROUND] Starting background processing for session {session_id}", file=sys.stderr)
    print(f"[BACKGROUND] Processing {len(file_paths)} files", file=sys.stderr)
    
    try:
        session.status = "extracting"
        session.progress = 10
        print(f"[BACKGROUND] Status set to 'extracting', progress: 10%", file=sys.stderr)
        
        # Initialize processor
        print(f"[BACKGROUND] Initializing DocumentProcessor...", file=sys.stderr)
        processor = DocumentProcessor()
        
        # Process documents
        async def progress_callback(progress_data):
            new_progress = 10 + int(progress_data["current"] / progress_data["total"] * 70)
            session.progress = new_progress
            session.status = progress_data["status"]
            print(f"[BACKGROUND] Progress update: {new_progress}% - {progress_data['status']}", file=sys.stderr)
        
        print(f"[BACKGROUND] Starting document processing...", file=sys.stderr)
        result = await processor.process_documents(file_paths, progress_callback)
        
        print(f"[BACKGROUND] Processing complete. Extracted {len(result.get('invoices', []))} invoices", file=sys.stderr)
        
        session.extracted_invoices = result.get("invoices", [])
        session.progress = 80
        session.status = "extracted"
        print(f"[BACKGROUND] Status set to 'extracted', progress: 80%", file=sys.stderr)
        
        # Generate initial Excel
        print(f"[BACKGROUND] Generating Excel report...", file=sys.stderr)
        generator = ExcelGenerator()
        excel_data, filename = generator.generate_invoice_sheet(session.extracted_invoices)
        session.excel_data = {
            "filename": filename,
            "size": len(excel_data),
            "data_preview": [inv for inv in session.extracted_invoices[:5]]  # First 5 for preview
        }
        
        session.progress = 100
        session.status = "completed"
        print(f"[BACKGROUND] ✓ Processing complete! Status: completed, progress: 100%", file=sys.stderr)
        
    except Exception as e:
        session.status = "error"
        session.error = str(e)
        print(f"[BACKGROUND] ✗ ERROR: {str(e)}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        session.progress = 0


@router.get("/progress/{session_id}")
async def get_progress(session_id: str):
    """Get processing progress for a session"""
    if session_id not in processing_jobs:
        print(f"[PROGRESS] Session not found: {session_id}", file=sys.stderr)
        raise HTTPException(status_code=404, detail="Session not found")
    
    session = processing_jobs[session_id]
    progress_data = {
        "session_id": session_id,
        "status": session.status,
        "progress": session.progress,
        "extracted_count": len(session.extracted_invoices),
        "error": session.error
    }
    print(f"[PROGRESS] Returning progress for {session_id}: {progress_data}", file=sys.stderr)
    return progress_data


@router.post("/upload-gstr2b/{session_id}")
async def upload_gstr2b(session_id: str, file: UploadFile = File(...)):
    """
    Upload GSTR2B data from Excel file for mismatch detection
    """
    if session_id not in processing_jobs:
        raise HTTPException(status_code=404, detail="Session not found")
    
    session = processing_jobs[session_id]
    
    try:
        # Validate file type
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are allowed")
        
        # Read the file
        content = await file.read()
        
        # Parse Excel file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_file.write(content)
            tmp_file_path = tmp_file.name
        
        try:
            gstr2b_data = _parse_gstr2b_excel(tmp_file_path)
            
            if not gstr2b_data or "invoices" not in gstr2b_data:
                raise HTTPException(status_code=400, detail="Invalid Excel format. Please ensure it contains invoice data.")
            
            # Add period from session if not in parsed data
            if not gstr2b_data.get("period"):
                gstr2b_data["period"] = session.month
            
            # Validate GSTR2B data
            processor = DocumentProcessor()
            validation_result = await processor.validate_gstr2b_data(gstr2b_data)
            
            if not validation_result["valid"]:
                raise HTTPException(status_code=400, detail=validation_result["message"])
            
            session.gstr2b_data = gstr2b_data
            session.status = "gstr2b_uploaded"
            
            return {
                "status": "success",
                "message": "GSTR2B data uploaded successfully",
                "session_id": session_id,
                "invoices_count": len(gstr2b_data.get("invoices", []))
            }
        finally:
            # Clean up temp file
            if os.path.exists(tmp_file_path):
                os.remove(tmp_file_path)
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing Excel file: {str(e)}")


@router.get("/govt-api/gstr2b")
async def fetch_gstr2b_from_govt(gstin: str, period: str):
    """
    Fetch GSTR2B from government API
    (Mock implementation - integrate with actual govt API)
    """
    try:
        # TODO: Integrate with actual GST government API
        # For now, return a structure that can be filled by the government API
        
        # This is a placeholder structure
        gstr2b_data = {
            "gstin": gstin,
            "period": period,
            "invoices": [
                # Will be populated by actual API
            ],
            "status": "fetched_from_govt"
        }
        
        return {
            "status": "success",
            "data": gstr2b_data,
            "message": "GSTR2B data fetched from government portal"
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/detect-mismatches/{session_id}")
async def detect_mismatches(session_id: str):
    """
    Run mismatch detection between extracted invoices and GSTR2B
    """
    if session_id not in processing_jobs:
        raise HTTPException(status_code=404, detail="Session not found")
    
    session = processing_jobs[session_id]
    
    if not session.extracted_invoices:
        raise HTTPException(status_code=400, detail="No extracted invoices available")
    
    if not session.gstr2b_data:
        raise HTTPException(status_code=400, detail="GSTR2B data not uploaded")
    
    try:
        session.status = "detecting_mismatches"
        session.progress = 0
        
        # Initialize detector
        detector = MismatchDetector()
        
        # Detect mismatches
        mismatch_results = detector.detect_mismatches(
            session.extracted_invoices,
            session.gstr2b_data
        )
        
        # Generate report card
        report_card = detector.generate_report_card(mismatch_results)
        
        session.mismatch_results = {
            "analysis": mismatch_results,
            "report_card": report_card
        }
        
        # Generate final Excel with highlighted mismatches
        generator = ExcelGenerator()
        excel_data, filename = generator.generate_mismatch_report_sheet(mismatch_results)
        
        session.excel_data = {
            "filename": filename,
            "size": len(excel_data),
            "type": "mismatch_report"
        }
        
        session.status = "mismatch_detection_completed"
        session.progress = 100
        
        return {
            "status": "success",
            "session_id": session_id,
            "report_card": report_card
        }
    
    except Exception as e:
        session.status = "error"
        session.error = str(e)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/reconcile-gstr2b/{session_id}")
async def reconcile_gstr2b(session_id: str):
    """
    Perform GSTR-2B reconciliation using MVP logic.
    Returns comprehensive reconciliation results with status, probable reasons, and actions.
    """
    if session_id not in processing_jobs:
        raise HTTPException(status_code=404, detail="Session not found")
    
    session = processing_jobs[session_id]
    
    if not session.extracted_invoices:
        raise HTTPException(status_code=400, detail="No extracted invoices available")
    
    if not session.gstr2b_data:
        raise HTTPException(status_code=400, detail="GSTR2B data not uploaded")
    
    try:
        session.status = "reconciling"
        session.progress = 0
        
        # Initialize detector with reconciliation engine
        detector = MismatchDetector()
        
        # Perform reconciliation
        reconciliation_result = detector.reconcile(
            session.extracted_invoices,
            session.gstr2b_data
        )
        
        session.mismatch_results = {
            "reconciliation": reconciliation_result
        }
        
        session.status = "reconciliation_completed"
        session.progress = 100
        
        return {
            "status": "success",
            "session_id": session_id,
            "reconciliation": reconciliation_result
        }
    
    except Exception as e:
        session.status = "error"
        session.error = str(e)
        print(f"Reconciliation error: {str(e)}", file=sys.stderr)
        raise HTTPException(status_code=500, detail=f"Reconciliation failed: {str(e)}")


@router.get("/session/{session_id}")
async def get_session_data(session_id: str):
    """Get complete session data including all processing results"""
    if session_id not in processing_jobs:
        raise HTTPException(status_code=404, detail="Session not found")
    
    session = processing_jobs[session_id]
    return session.to_dict()


@router.get("/download-excel/{session_id}")
async def download_excel(session_id: str):
    """Download the generated Excel file"""
    if session_id not in processing_jobs:
        raise HTTPException(status_code=404, detail="Session not found")
    
    session = processing_jobs[session_id]
    
    if not session.excel_data:
        raise HTTPException(status_code=400, detail="Excel file not generated yet")
    
    try:
        # Regenerate Excel on download
        generator = ExcelGenerator()
        
        if session.mismatch_results:
            excel_data, filename = generator.generate_mismatch_report_sheet(
                session.mismatch_results["analysis"]
            )
        else:
            excel_data, filename = generator.generate_invoice_sheet(session.extracted_invoices)
        
        return StreamingResponse(
            iter([excel_data]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/update-excel/{session_id}")
async def update_excel(session_id: str, updates: Dict):
    """
    Update Excel data with manual edits
    (Store edits and regenerate Excel)
    """
    if session_id not in processing_jobs:
        raise HTTPException(status_code=404, detail="Session not found")
    
    session = processing_jobs[session_id]
    
    try:
        # Apply updates to extracted invoices or GSTR2B data
        if "invoices" in updates:
            session.extracted_invoices = updates["invoices"]
        
        # Regenerate mismatch detection if needed
        if session.gstr2b_data and session.extracted_invoices:
            detector = MismatchDetector()
            mismatch_results = detector.detect_mismatches(
                session.extracted_invoices,
                session.gstr2b_data
            )
            
            session.mismatch_results = {
                "analysis": mismatch_results,
                "report_card": detector.generate_report_card(mismatch_results)
            }
        
        return {
            "status": "success",
            "message": "Excel data updated",
            "session_id": session_id
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/session/{session_id}")
async def delete_session(session_id: str):
    """Delete a processing session"""
    if session_id not in processing_jobs:
        raise HTTPException(status_code=404, detail="Session not found")
    
    del processing_jobs[session_id]
    
    return {
        "status": "success",
        "message": "Session deleted"
    }
