from typing import Dict, List, Tuple
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

class ExcelGenerator:
    """Handles generation and manipulation of Excel sheets"""
    
    def __init__(self):
        self.highlight_color = "FFFF00"  # Yellow for mismatches
        self.error_color = "FF0000"  # Red for errors
        self.match_color = "00B050"  # Green for matches
    
    def generate_invoice_sheet(self, invoices: List[Dict], title: str = "Extracted Invoices") -> Tuple[bytes, str]:
        """
        Generate Excel sheet from extracted invoice data
        
        Returns:
            Tuple of (excel_bytes, filename)
        """
        df = self._prepare_dataframe(invoices)
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Invoices"
        
        # Write headers
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Adjust column widths
        for col_num, column_title in enumerate(df.columns, 1):
            column_letter = get_column_letter(col_num)
            max_length = len(str(column_title)) + 2
            worksheet.column_dimensions[column_letter].width = min(max_length + 5, 50)
        
        # Save to bytes
        excel_bytes = io.BytesIO()
        workbook.save(excel_bytes)
        excel_bytes.seek(0)
        
        return excel_bytes.getvalue(), "invoices.xlsx"
    
    def generate_mismatch_report_sheet(self, mismatch_data: Dict) -> Tuple[bytes, str]:
        """
        Generate Excel sheet with mismatch analysis and highlighted differences
        
        Returns:
            Tuple of (excel_bytes, filename)
        """
        workbook = Workbook()
        
        # Sheet 1: Summary
        summary_ws = workbook.active
        summary_ws.title = "Summary"
        self._write_summary_sheet(summary_ws, mismatch_data)
        
        # Sheet 2: Matched Invoices
        if mismatch_data["matched_pairs"]:
            matched_ws = workbook.create_sheet("Matched")
            self._write_matched_sheet(matched_ws, mismatch_data["matched_pairs"])
        
        # Sheet 3: Mismatches
        if mismatch_data["mismatches"]:
            mismatch_ws = workbook.create_sheet("Mismatches")
            self._write_mismatches_sheet(mismatch_ws, mismatch_data["mismatches"])
        
        # Sheet 4: Unmatched Extracted
        if mismatch_data["unmatched_extracted"]:
            unmatched_ext_ws = workbook.create_sheet("Unmatched Extracted")
            self._write_unmatched_extracted_sheet(unmatched_ext_ws, mismatch_data["unmatched_extracted"])
        
        # Sheet 5: Unmatched GSTR2B
        if mismatch_data["unmatched_gstr2b"]:
            unmatched_gstr_ws = workbook.create_sheet("Unmatched GSTR2B")
            self._write_unmatched_gstr2b_sheet(unmatched_gstr_ws, mismatch_data["unmatched_gstr2b"])
        
        # Save to bytes
        excel_bytes = io.BytesIO()
        workbook.save(excel_bytes)
        excel_bytes.seek(0)
        
        return excel_bytes.getvalue(), "mismatch_report.xlsx"
    
    def _write_summary_sheet(self, worksheet, mismatch_data: Dict):
        """Write summary information to worksheet"""
        summary = mismatch_data["summary"]
        
        worksheet.column_dimensions["A"].width = 40
        worksheet.column_dimensions["B"].width = 20
        
        headers = ["Metric", "Value"]
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        data = [
            ["Total Invoices Extracted", summary["total_extracted"]],
            ["Total Invoices in GSTR2B", summary["total_gstr2b"]],
            ["Successfully Matched", summary["matched"]],
            ["Discrepancies Found", summary["mismatch_count"]],
            ["Missing from GSTR2B", summary["unmatched_extracted"]],
            ["Extra in GSTR2B", summary["unmatched_gstr2b"]],
            ["Match Rate (%)", round((summary["matched"] / summary["total_extracted"] * 100) if summary["total_extracted"] > 0 else 0, 2)]
        ]
        
        for row_num, row_data in enumerate(data, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
                if col_num == 2:
                    cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="left", vertical="center")
    
    def _write_matched_sheet(self, worksheet, matched_pairs: List[Dict]):
        """Write matched invoices to worksheet"""
        headers = ["Invoice #", "Date", "GSTIN", "Extracted Amount", "GSTR2B Amount", "Match Score", "Issues"]
        
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for row_num, pair in enumerate(matched_pairs, 2):
            ext = pair["extracted"]
            gstr = pair["gstr2b"]
            
            row_data = [
                ext.get("invoice_number", "N/A"),
                ext.get("invoice_date", "N/A"),
                ext.get("gstin", "N/A"),
                ext.get("total_amount", 0),
                gstr.get("total_amount", 0),
                round(pair["match_score"], 3),
                "; ".join(pair["mismatches"]) if pair["mismatches"] else "No issues"
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                # Highlight rows with mismatches
                if pair["mismatches"]:
                    cell.fill = PatternFill(start_color=self.highlight_color, end_color=self.highlight_color, fill_type="solid")
        
        # Adjust column widths
        for col_num in range(1, len(headers) + 1):
            worksheet.column_dimensions[get_column_letter(col_num)].width = 20
    
    def _write_mismatches_sheet(self, worksheet, mismatches: List[Dict]):
        """Write mismatch details to worksheet"""
        headers = ["Invoice #", "Match Score", "Issues"]
        
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        
        for row_num, mismatch in enumerate(mismatches, 2):
            row_data = [
                mismatch["invoice_number"],
                round(mismatch["match_score"], 3),
                "\n".join(mismatch["issues"])
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.fill = PatternFill(start_color=self.highlight_color, end_color=self.highlight_color, fill_type="solid")
        
        # Adjust column widths
        worksheet.column_dimensions["A"].width = 20
        worksheet.column_dimensions["B"].width = 15
        worksheet.column_dimensions["C"].width = 50
    
    def _write_unmatched_extracted_sheet(self, worksheet, unmatched: List[Dict]):
        """Write unmatched extracted invoices"""
        headers = ["File", "Invoice #", "Amount", "Reason"]
        
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
        
        for row_num, item in enumerate(unmatched, 2):
            inv = item["invoice"]
            row_data = [
                inv.get("file", "N/A"),
                inv.get("invoice_number", "N/A"),
                inv.get("total_amount", 0),
                item["reason"]
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.fill = PatternFill(start_color=self.error_color, end_color=self.error_color, fill_type="solid")
                cell.font = Font(color="FFFFFF")
        
        for col_num in range(1, len(headers) + 1):
            worksheet.column_dimensions[get_column_letter(col_num)].width = 20
    
    def _write_unmatched_gstr2b_sheet(self, worksheet, unmatched: List[Dict]):
        """Write unmatched GSTR2B invoices"""
        headers = ["Invoice #", "Date", "GSTIN", "Amount", "Status"]
        
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
        
        for row_num, inv in enumerate(unmatched, 2):
            row_data = [
                inv.get("invoice_number", "N/A"),
                inv.get("invoice_date", "N/A"),
                inv.get("gstin", "N/A"),
                inv.get("total_amount", 0),
                "Not found in extracted invoices"
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.fill = PatternFill(start_color=self.error_color, end_color=self.error_color, fill_type="solid")
                cell.font = Font(color="FFFFFF")
        
        for col_num in range(1, len(headers) + 1):
            worksheet.column_dimensions[get_column_letter(col_num)].width = 20
    
    def _prepare_dataframe(self, invoices: List[Dict]) -> pd.DataFrame:
        """Prepare dataframe from invoice list with all GST and reconciliation fields"""
        data = []
        
        for inv in invoices:
            if inv.get("status") == "error":
                data.append({
                    "File": inv.get("file", "N/A"),
                    "Supplier GSTIN": "ERROR",
                    "Invoice No": "ERROR",
                    "Invoice Date": "ERROR",
                    "Document Type": "ERROR",
                    "Taxable Value": 0,
                    "CGST": 0,
                    "SGST": 0,
                    "IGST": 0,
                    "Total Amount": 0,
                    "Expense Category": inv.get("error", "Unknown error"),
                    "ITC Eligibility": "N/A",
                    "GSTR2B Section": "N/A",
                    "Status": "error"
                })
            else:
                data.append({
                    "File": inv.get("file", "N/A"),
                    "Supplier GSTIN": inv.get("supplier_gstin", "N/A"),
                    "Invoice No": inv.get("invoice_number", "N/A"),
                    "Invoice Date": inv.get("invoice_date", "N/A"),
                    "Document Type": inv.get("document_type", "Invoice"),
                    "Taxable Value": inv.get("taxable_value", 0),
                    "CGST": inv.get("cgst", 0),
                    "SGST": inv.get("sgst", 0),
                    "IGST": inv.get("igst", 0),
                    "Total Amount": inv.get("total_amount", 0),
                    "Expense Category": inv.get("expense_category", "N/A"),
                    "ITC Eligibility": inv.get("itc_eligibility", "N/A"),
                    "GSTR2B Section": inv.get("gstr2b_section", "N/A"),
                    "Status": inv.get("status", "unknown")
                })
        
        return pd.DataFrame(data)
